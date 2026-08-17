"""
test_parse.py - checks parse.py's five sheet readers against tiny fixture workbooks.

Each test builds a miniature .xlsx in memory (via openpyxl) that mimics the exact row
layout of the real ABS/JSA file it stands in for - same header rows, same staircase or
flat structure, same footer noise - so the tests catch a row-offset regression the way
the real bug (growth percentages read from the wrong column, fixed during the original
build) would have been caught if it had existed here first.

Run: uv run pytest tests/test_parse.py -q
"""

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import parse  # noqa: E402


def _save(wb, tmp_path, name):
    path = tmp_path / name
    wb.save(path)
    return path


def test_load_anzsco_structure(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table 4"
    # 9 blank/header rows to match the real file's staircase header, data starts row 10
    for _ in range(9):
        ws.append([None] * 8)
    ws.append([1, "Managers", None, None, None, None, None, None])
    ws.append([None, 11, "Chief Executives, General Managers and Legislators", None, None, None, None, None])
    ws.append([None, None, 111, "Chief Executives, General Managers and Legislators", None, None, None, None])
    ws.append([None, None, None, 1111, "Chief Executives and Managing Directors", 1, None, None])
    ws.append([None, None, None, 1112, "General Managers", 1, None, None])
    ws.append([2, "Professionals", None, None, None, None, None, None])
    ws.append([None, None, None, 2211, "Accountants", 1, None, None])
    ws.append([None] * 8)
    ws.append(["© Commonwealth of Australia", None, None, None, None, None, None, None])
    ws.append(["Crown Copyright ©", None, None, None, None, None, None, None])
    path = _save(wb, tmp_path, "structure.xlsx")

    occupations = parse.load_anzsco_structure(path)

    assert set(occupations.keys()) == {1111, 1112, 2211}
    assert occupations[1111].title == "Chief Executives and Managing Directors"
    assert occupations[1111].major_group_code == 1
    assert occupations[1111].major_group == "Managers"
    assert occupations[1111].skill_level == "1"
    assert occupations[2211].major_group == "Professionals"


def _base_occupations():
    return {
        1111: parse.Occupation(anzsco_code=1111, title="Chief Executives and Managing Directors"),
        2211: parse.Occupation(anzsco_code=2211, title="Accountants"),
    }


def test_load_employment_keeps_only_latest_quarter_and_sums_sex_and_state(tmp_path):
    import datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data 1"
    for _ in range(4):
        ws.append([None] * 6)  # 4 header/title rows, data starts row 5
    older = datetime.datetime(2025, 11, 1)
    latest = datetime.datetime(2026, 2, 1)
    # older quarter: should be dropped once the newer quarter appears
    ws.append([older, "Males", "New South Wales", "1111 Chief Executives and Managing Directors", 999.0, 0])
    # latest quarter, code 1111: two sexes x two states = four rows to sum
    ws.append([latest, "Males", "New South Wales", "1111 Chief Executives and Managing Directors", 10.0, 0])
    ws.append([latest, "Females", "New South Wales", "1111 Chief Executives and Managing Directors", 5.0, 0])
    ws.append([latest, "Males", "Victoria", "1111 Chief Executives and Managing Directors", 8.0, 0])
    ws.append([latest, "Females", "Victoria", "1111 Chief Executives and Managing Directors", 4.0, 0])
    ws.append([latest, "Males", "New South Wales", "2211 Accountants", 20.0, 0])
    path = _save(wb, tmp_path, "eq08.xlsx")

    occupations = _base_occupations()
    matched, period = parse.load_employment(occupations, path)

    assert matched == 2
    assert period == "February 2026"
    assert occupations[1111].employment_thousands == pytest.approx(27.0)
    assert occupations[2211].employment_thousands == pytest.approx(20.0)


def test_load_earnings_leaves_suppressed_cells_as_none(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table_1"
    for _ in range(6):
        ws.append([None] * 4)  # data starts row 7
    ws.append(["1111 Chief executives and managing directors ", 2775.6, 3569.7, 2962.8])
    ws.append(["2211 Accountants ", None, None, None])  # suppressed
    path = _save(wb, tmp_path, "earnings.xlsx")

    occupations = _base_occupations()
    matched = parse.load_earnings(occupations, path)

    assert matched == 1
    assert occupations[1111].avg_weekly_earnings == pytest.approx(2962.8)
    assert occupations[2211].avg_weekly_earnings is None


def test_load_projections_reads_percentage_columns_not_level_columns(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table_6 Occupation Unit Group"
    for _ in range(9):
        ws.append([None] * 12)  # data starts row 10
    # a not-further-defined placeholder row: must be skipped
    ws.append([4, "Y", 1000, "Managers nfd", "-", 6.15, 6.73, 7.20, 0.58, 0.0940, 1.05, 0.1698])
    # a real unit group row: index 9 is the 5yr %, index 11 is the 10yr % (not 8 or 10,
    # which are the level-change columns in '000 - the original bug used those instead)
    ws.append([4, "N", 1111, "Chief Executives and Managing Directors", 1,
               61.19, 65.39, 69.82, 4.19, 0.0685, 8.62, 0.1409])
    # wrong level: must be skipped
    ws.append([2, "N", 11, "Chief Executives, General Managers and Legislators", "-",
               135.5, 144.5, 154.3, 8.99, 0.0663, 18.76, 0.1385])
    path = _save(wb, tmp_path, "projections.xlsx")

    occupations = _base_occupations()
    matched = parse.load_projections(occupations, path)

    assert matched == 1
    assert occupations[1111].growth_5y_pct == pytest.approx(6.85)
    assert occupations[1111].growth_10y_pct == pytest.approx(14.09)
    assert occupations[2211].growth_5y_pct is None


def test_load_shortage(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2025 Unit group Shortage List"
    for _ in range(8):
        ws.append([None] * 3)  # data starts row 9
    ws.append([1111, "Chief Executives and Managing Directors", "NS"])
    path = _save(wb, tmp_path, "shortage.xlsx")

    occupations = _base_occupations()
    matched = parse.load_shortage(occupations, path)

    assert matched == 1
    assert occupations[1111].shortage_rating == "NS"
    assert occupations[2211].shortage_rating is None
