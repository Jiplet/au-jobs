#!/usr/bin/env python3
"""
parse.py - turn the raw ABS/JSA spreadsheets into one tidy occupation table.

What it does:
    Reads the five files fetch.py downloaded into data/raw/ and merges them on ANZSCO
    4-digit occupation unit group code into one row per occupation:
      - abs_anzsco_structure.xlsx  -> the master list of unit group codes, titles, and
                                       which major group (1-digit) each belongs to
      - abs_eq08.xlsx              -> employment (headcount), latest quarter, Australia
                                       total, persons (summed from the Male/Female by
                                       state rows the file actually publishes)
      - abs_earnings.xlsx          -> average weekly total cash earnings, persons
                                       (NOT a median - see note below)
      - jsa_employment_projections.xlsx -> projected employment growth, 5-year and
                                       10-year horizons
      - jsa_unit_group_shortage.xlsx -> national skills shortage rating

    A note on earnings: the brief for this project asked for median earnings. ABS
    Employee Earnings and Hours only publishes an AVERAGE (mean) at ANZSCO unit group
    level, not a median. Shipping the real average under an honest column name beats
    either faking a median or dropping the layer, so the CSV column is
    avg_weekly_earnings and every doc that mentions it says "average", not "median".

    A note on task descriptions: neither ABS nor JSA publish a bulk file of occupation
    task lists or lead statements (ABS only has them as one HTML page per unit group,
    ~360 pages, not worth scraping for this build). So tasks_text ships empty and
    score.py asks the model to infer likely tasks from the title and major group before
    scoring - see prompts/ai-exposure.md. This is documented as a weaker input, not
    hidden.

Inputs:
    data/raw/*.xlsx (from fetch.py).

Outputs:
    data/occupations.csv - one row per ANZSCO 4-digit unit group.
    data/parse-report.md - row counts and coverage gaps per layer, written fresh each run.

How to run it:
    uv run python fetch.py     # if you have not already
    uv run python parse.py
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent
RAW_DIR = REPO_ROOT / "data" / "raw"
OUT_CSV = REPO_ROOT / "data" / "occupations.csv"
OUT_REPORT = REPO_ROOT / "data" / "parse-report.md"

CODE_TITLE_RE = re.compile(r"^\s*(\d{4})\s+(.+?)\s*$")


@dataclass
class Occupation:
    anzsco_code: int
    title: str
    major_group_code: int | None = None
    major_group: str | None = None
    skill_level: str | None = None
    employment_thousands: float | None = None
    avg_weekly_earnings: float | None = None
    growth_5y_pct: float | None = None
    growth_10y_pct: float | None = None
    shortage_rating: str | None = None
    tasks_text: str = ""


def load_anzsco_structure(path: Path | None = None) -> dict[int, Occupation]:
    """Table 4 of the ANZSCO structure file is a staircase: the first non-empty of the
    first four columns tells you the level (major/sub-major/minor/unit group), and the
    code sits in that column with its title one column to the right. Unit group rows
    also carry a skill level in the sixth column.

    `path` defaults to the real downloaded file; tests pass a tiny fixture instead."""
    wb = openpyxl.load_workbook(path or RAW_DIR / "abs_anzsco_structure.xlsx", read_only=True, data_only=True)
    ws = wb["Table 4"]

    occupations: dict[int, Occupation] = {}
    current_major_code: int | None = None
    current_major_title: str | None = None

    for row in ws.iter_rows(min_row=10, values_only=True):
        level = None
        for idx in range(4):
            if row[idx] is not None:
                level = idx
                break
        if level is None:
            continue

        if level == 0:
            if not isinstance(row[0], (int, float)):
                break  # footer row (copyright notice), end of the real table
            current_major_code = int(row[0])
            current_major_title = row[1]
        elif level == 3:
            code = int(row[3])
            occupations[code] = Occupation(
                anzsco_code=code,
                title=str(row[4]).strip(),
                major_group_code=current_major_code,
                major_group=current_major_title,
                skill_level=str(row[5]) if row[5] is not None else None,
            )
        # levels 1 (sub-major) and 2 (minor) are not needed at unit-group grain

    wb.close()
    return occupations


def load_employment(occupations: dict[int, Occupation], path: Path | None = None) -> tuple[int, str]:
    """EQ08's flat data has no pre-aggregated Australia/Persons rows - it publishes
    Male and Female counts for each of the 8 states/territories separately, one row
    per (quarter, sex, state, occupation). We sum across sex and state for the most
    recent quarter to get Australia total, persons. One linear pass: whenever a later
    quarter than the one we are accumulating shows up, the accumulator resets, so by
    the end it holds only the latest quarter regardless of how many quarters preceded
    it in the file."""
    wb = openpyxl.load_workbook(path or RAW_DIR / "abs_eq08.xlsx", read_only=True, data_only=True)
    ws = wb["Data 1"]

    totals: dict[int, float] = defaultdict(float)
    current_date = None
    matched = 0
    unmatched_codes: set[str] = set()

    for row in ws.iter_rows(min_row=5, values_only=True):
        quarter, sex, state, occ_label, employed_total = row[0], row[1], row[2], row[3], row[4]
        if quarter is None or occ_label is None or employed_total is None:
            continue
        if current_date is None or quarter > current_date:
            current_date = quarter
            totals.clear()
        if quarter != current_date:
            continue
        m = CODE_TITLE_RE.match(str(occ_label))
        if not m:
            continue
        code = int(m.group(1))
        totals[code] += float(employed_total)

    wb.close()

    for code, total in totals.items():
        if code in occupations:
            occupations[code].employment_thousands = round(total, 3)
            matched += 1
        else:
            unmatched_codes.add(str(code))

    if unmatched_codes:
        print(f"  note: {len(unmatched_codes)} EQ08 occupation code(s) not found in the "
              f"ANZSCO structure master list, skipped: {sorted(unmatched_codes)}")

    period = current_date.strftime("%B %Y") if current_date else "unknown"
    return matched, period


def load_earnings(occupations: dict[int, Occupation], path: Path | None = None) -> int:
    """Table_1: one row per unit group, code and title packed into column A as e.g.
    '1111 Chief executives and managing directors '. Column D (index 3) is the
    Persons average weekly total cash earnings. Suppressed cells (too few respondents)
    are blank and stay None rather than being coerced to zero."""
    wb = openpyxl.load_workbook(path or RAW_DIR / "abs_earnings.xlsx", read_only=True, data_only=True)
    ws = wb["Table_1"]

    matched = 0
    for row in ws.iter_rows(min_row=7, values_only=True):
        label, persons_earnings = row[0], row[3]
        if label is None:
            continue
        m = CODE_TITLE_RE.match(str(label))
        if not m:
            continue
        code = int(m.group(1))
        if code in occupations and persons_earnings is not None:
            occupations[code].avg_weekly_earnings = round(float(persons_earnings), 2)
            matched += 1

    wb.close()
    return matched


def load_projections(occupations: dict[int, Occupation], path: Path | None = None) -> int:
    """Table_6: every classification level is mixed into one sheet, flagged by an
    'Occupation Level' column (4 = unit group) and an 'NFD Indicator' ('N' = a real,
    fully-described unit group; 'Y' = a not-further-defined placeholder row that exists
    for totals and should be skipped)."""
    wb = openpyxl.load_workbook(path or RAW_DIR / "jsa_employment_projections.xlsx", read_only=True, data_only=True)
    ws = wb["Table_6 Occupation Unit Group"]

    matched = 0
    for row in ws.iter_rows(min_row=10, values_only=True):
        level, nfd, code = row[0], row[1], row[2]
        if level != 4 or nfd != "N" or code is None:
            continue
        code = int(code)
        growth_5y_pct, growth_10y_pct = row[9], row[11]
        if code in occupations:
            occupations[code].growth_5y_pct = round(float(growth_5y_pct) * 100, 2) if growth_5y_pct is not None else None
            occupations[code].growth_10y_pct = round(float(growth_10y_pct) * 100, 2) if growth_10y_pct is not None else None
            matched += 1

    wb.close()
    return matched


def load_shortage(occupations: dict[int, Occupation], path: Path | None = None) -> int:
    """One clean row per unit group: code, title, then the national rating
    (NS/S/R/M = No Shortage / Shortage / Regional Shortage / Metropolitan Shortage)."""
    wb = openpyxl.load_workbook(path or RAW_DIR / "jsa_unit_group_shortage.xlsx", read_only=True, data_only=True)
    ws = wb["2025 Unit group Shortage List"]

    matched = 0
    for row in ws.iter_rows(min_row=9, values_only=True):
        code, rating = row[0], row[2]
        if code is None:
            continue
        code = int(code)
        if code in occupations and rating is not None:
            occupations[code].shortage_rating = str(rating).strip()
            matched += 1

    wb.close()
    return matched


def write_csv(occupations: dict[int, Occupation]) -> None:
    import csv

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "anzsco_code", "title", "major_group_code", "major_group", "skill_level",
        "employment_thousands", "avg_weekly_earnings", "growth_5y_pct", "growth_10y_pct",
        "shortage_rating", "tasks_text",
    ]
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for code in sorted(occupations):
            writer.writerow(vars(occupations[code]))


def write_report(occupations: dict[int, Occupation], employment_matched: int, employment_period: str,
                  earnings_matched: int, projections_matched: int, shortage_matched: int) -> None:
    total = len(occupations)

    def pct(n: int) -> str:
        return f"{n}/{total} ({n / total * 100:.0f}%)" if total else f"{n}/0"

    lines = [
        "# Parse report",
        "",
        f"Total ANZSCO 4-digit unit groups (from the ABS ANZSCO 2022 structure file): {total}",
        "",
        "## Coverage per layer",
        "",
        "| Layer | Matched | Notes |",
        "|---|---|---|",
        f"| Employment ({employment_period}) | {pct(employment_matched)} | Australia total, persons, summed from Male+Female across all 8 states/territories - the source file has no pre-aggregated total. |",
        f"| Average weekly earnings | {pct(earnings_matched)} | This is a MEAN, not a median - ABS does not publish a median at unit group level. Many small occupations are suppressed by ABS (too few survey respondents) and are correctly blank, not zero. |",
        f"| Projected employment growth | {pct(projections_matched)} | 5-year (to May 2030) and 10-year (to May 2035), JSA Employment Projections. |",
        f"| Skills shortage rating | {pct(shortage_matched)} | National rating, JSA Occupation Shortage List cycle 2025, pre-aggregated to unit group by JSA. |",
        f"| Task descriptions | 0/{total} (0%) | Not available as a bulk file from either ABS or JSA (ABS only publishes lead statements as ~360 individual HTML pages, one per unit group). score.py asks the LLM to infer likely tasks from the title and major group before scoring - see prompts/ai-exposure.md. This is a real gap, documented rather than papered over. |",
        "",
        "## What this means for the map",
        "",
        "Every occupation has a code, title, and major group. Employment coverage should be "
        "at or near 100% since it comes from the same classification vintage as the "
        "structure file. Earnings, growth, and shortage will each have some gaps: earnings "
        "from ABS suppression of small occupations, growth and shortage from vintage "
        "mismatches between JSA's occupation list and the ABS 2022 structure. The site's "
        "legend marks a layer 'not available for this occupation' rather than plotting a "
        "zero or a made-up value.",
    ]
    OUT_REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    print("Loading ANZSCO structure (master list)...")
    occupations = load_anzsco_structure()
    print(f"  {len(occupations)} unit groups")

    print("Loading employment (EQ08, this is the big file, may take ~30s)...")
    employment_matched, employment_period = load_employment(occupations)
    print(f"  {employment_matched} matched, latest quarter: {employment_period}")

    print("Loading earnings...")
    earnings_matched = load_earnings(occupations)
    print(f"  {earnings_matched} matched")

    print("Loading employment projections...")
    projections_matched = load_projections(occupations)
    print(f"  {projections_matched} matched")

    print("Loading skills shortage ratings...")
    shortage_matched = load_shortage(occupations)
    print(f"  {shortage_matched} matched")

    write_csv(occupations)
    print(f"Wrote {OUT_CSV.relative_to(REPO_ROOT)}")

    write_report(occupations, employment_matched, employment_period,
                 earnings_matched, projections_matched, shortage_matched)
    print(f"Wrote {OUT_REPORT.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
